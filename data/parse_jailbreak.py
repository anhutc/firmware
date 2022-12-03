
from typing import List, Dict
import openpyxl
import json

wb: openpyxl.workbook.Workbook = openpyxl.load_workbook('jailbreak.xlsx')
colorTable: Dict[int, str] = {
    5: 'TETHERED',
    7: 'SEMI-UNTETHERED',
    8: 'SEMI-TETHERED',
    9: 'UNTETHERED'
}


def getVersionName(sheet: str, version: str) -> str:
    if sheet == 'iOS' and int(version.split('.')[0]) < 4:
        return 'iPhone OS ' + version
    else:
        return sheet + ' ' + version


sheets: List[str] = ['iOS', 'iPadOS', 'tvOS', 'watchOS', 'audioOS']
data: Dict[str, dict] = {}
prefix: str = ''
for sheet in sheets:
    print('Dealing with sheet %s...' % sheet)
    ws: openpyxl.worksheet.worksheet.Worksheet = wb[sheet]
    max_col: int = ws.max_column
    for i in range(1, max_col):
        # Get product name
        product: str = str(ws[1][i].value)
        # Firmware arraylist
        arr = {}
        if ws[1][i].font.b:
            arr = data.get(product, {})
            print('> %s...' % product)
            prefix = (product.split(
                '(')[0].strip() if product[-1] == ')' else product) + ' '
        else:
            product = prefix + product
            arr = data.get(product, {})
            print('> %s...' % product)

        for j in range(2, ws.max_row + 1):
            obj = {}
            index = ws[j][i].fill.start_color.index
            text: str = str(ws[j][i].value)
            version: str = str(ws[j][0].value).split(' ')[-1]
            version = getVersionName(sheet, version)
            if isinstance(index, int):
                if index != 0:
                    obj['type'] = colorTable[index]
                    obj['tool'] = text
                    if ws[j][i].comment is not None:
                        obj['comment'] = ws[j][i].comment.text
                    arr[version] = obj
            else:
                # No available jailbreak
                obj['type'] = 'NONE'
                arr[version] = obj
                if ws[j][i].comment is not None:
                    obj['comment'] = ws[j][i].comment.text
        data[product] = arr

with open('jailbreak.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False)
